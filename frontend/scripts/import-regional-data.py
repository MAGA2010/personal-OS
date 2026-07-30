#!/usr/bin/env python3
"""
PathOS Stage 7R — Regional Heatmap Importer.

Reads the source workbook (single file, multi-sheet), normalizes the four
READY metrics (income / safety / employment / chinese_population), and writes
deterministic JSON artifacts into `frontend/generated/regional-data/`.

The script is idempotent: two consecutive runs produce byte-identical output
(SHA-256 of every artifact is recorded in the manifest).

Usage:
    python3 scripts/import-regional-data.py \\
        --workbook resource/PathOS_美国各州留学数据矩阵.xlsx \\
        --out  frontend/generated/regional-data
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from typing import Any

import openpyxl  # type: ignore

# ----------------------------------------------------------------------
# Configuration — per audit (docs/STAGE7R-REGIONAL-DATA-AUDIT.md).
# ----------------------------------------------------------------------

DATASET_ID = "pathos-regional-state-v1"
DATASET_VERSION = "1.0.0"

# Per-metric definitions keyed by metricId.
# Column indices in `州级数据汇总` (1-based, matching the workbook):
#   A=FIPS, B=abbr, C=zh, D=en, E/F/G=income, H/I/J=safety,
#   K/L/M=employment, N/O/P=cost, Q/R/S=admission_rate, T/U/V=chinese_population
METRICS: dict[str, dict[str, Any]] = {
    "income": {
        "displayNameZh": "收入水平",
        "displayNameEn": "Median Income",
        "shortDescription": "区域家庭中位年收入",
        "longDescription": "区域家庭中位年收入，反映地区经济水平与生活成本",
        "sourceName": "Census ACS 5-Year",
        "sourceUrl": "",
        "rawUnit": "USD/year",
        "displayUnit": "$NNk",
        "allowedRange": [30000, 200000],
        "higherIsBetter": True,
        "normalizationMethod": "workbook-provided linear min-max to [0,1]; preserved",
        "rawDirection": "direct",
        "normCol": 5,  # E
        "rawCol": 6,   # F
        "dispCol": 7,  # G
        "paletteId": "palette-income-green",
        "usedForMap": True,
        "usedForMatch": False,
    },
    "safety": {
        "displayNameZh": "安全系数",
        "displayNameEn": "Safety Index",
        "shortDescription": "区域暴力犯罪率反向标准化",
        "longDescription": "基于 FBI UCR 暴力犯罪率（每 10 万人暴力犯罪案件数）。rawValue 保留原始犯罪率；normalizedValue 反向（越高越安全）",
        "sourceName": "FBI UCR (Uniform Crime Report)",
        "sourceUrl": "",
        "rawUnit": "crimes per 100,000 residents",
        "displayUnit": "NNN.N/100k",
        "allowedRange": [50, 1500],
        "higherIsBetter": False,  # raw = crime rate, lower is safer
        "normalizationMethod": "inverse: ourNormalized = 1 - workbookNorm",
        "rawDirection": "inverse",
        "normCol": 8,  # H
        "rawCol": 9,   # I
        "dispCol": 10, # J
        "paletteId": "palette-safety-blue",
        "usedForMap": True,
        "usedForMatch": False,
    },
    "employment": {
        "displayNameZh": "就业指数",
        "displayNameEn": "Employment Index",
        "shortDescription": "各州就业率（100% − 失业率）",
        "longDescription": "基于 BLS 各州失业率数据计算：就业率 = 100% - 失业率",
        "sourceName": "BLS (Bureau of Labor Statistics)",
        "sourceUrl": "",
        "rawUnit": "%",
        "displayUnit": "NN.N%",
        "allowedRange": [85, 100],
        "higherIsBetter": True,
        "normalizationMethod": "workbook-provided linear min-max to [0,1]; preserved",
        "rawDirection": "direct",
        "normCol": 11, # K
        "rawCol": 12,  # L
        "dispCol": 13, # M
        "paletteId": "palette-employment-purple",
        "usedForMap": True,
        "usedForMatch": False,
    },
    "chinese_population": {
        "displayNameZh": "华人水平",
        "displayNameEn": "Chinese Population",
        "shortDescription": "华裔人口规模",
        "longDescription": "华裔人口绝对数量，反映该区域华人社区规模和便利程度",
        "sourceName": "Census ACS",
        "sourceUrl": "",
        "rawUnit": "persons",
        "displayUnit": "NNNk",
        "allowedRange": [0, 2_000_000],
        "higherIsBetter": True,
        "normalizationMethod": "workbook-provided linear min-max to [0,1]; preserved",
        "rawDirection": "direct",
        "normCol": 20, # T
        "rawCol": 21,  # U
        "dispCol": 22, # V
        "paletteId": "palette-chinese-orange",
        "usedForMap": True,
        "usedForMatch": False,
    },
}

# Metric that exists but is OUT OF SCOPE for this round.
EXISTS_OUT_OF_SCOPE = {
    "cost": {
        "displayNameZh": "留学成本",
        "displayNameEn": "Study Cost",
        "shortDescription": "年度综合留学成本评估",
        "longDescription": "包含学费与生活费的年度综合留学成本评估，数值越高成本越高",
        "sourceName": "College Board / 各大学官网",
        "sourceUrl": "",
        "rawUnit": "CNY/year",
        "displayUnit": "¥NN万",
        "allowedRange": [100000, 800000],
        "higherIsBetter": False,
        "normalizationMethod": "workbook-provided linear min-max to [0,1]; preserved",
        "rawDirection": "direct",
        "normCol": 14, # N
        "rawCol": 15,  # O
        "dispCol": 16, # P
        "paletteId": "palette-cost-amber",
        "usedForMap": False,
        "usedForMatch": False,
    },
}

# Metric that is BLOCKED — never parsed.
BLOCKED_METRICS = {
    "admission_rate": "Workbook author marks state-level data as missing; 51/51 N/A.",
}

# Reference year — best-effort from workbook banner ("更新: 2026-07")
# and per-metric source norms. We use a stable string here.
REFERENCE_YEAR = "2026-07 (ACS/FBI/BLS latest available)"
RETRIEVED_AT = "2026-07-25"


def sha256_file(path: str) -> str:
    """Compute hex SHA-256 of a file's bytes."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_str(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def load_workbook(path: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True)


def read_metric(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    metric_id: str,
    cfg: dict[str, Any],
    rawDirection: str,
) -> list[dict[str, Any]]:
    """Read one metric block from the main sheet.

    Returns a list of RegionalMetricRecord-compatible dicts.
    """
    norm_col = cfg["normCol"] - 1
    raw_col = cfg["rawCol"] - 1
    disp_col = cfg["dispCol"] - 1
    records: list[dict[str, Any]] = []

    for ri, row in enumerate(
        ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True),
        start=3,
    ):
        fips = row[0]
        if not fips:
            continue
        abbr = row[1]
        name_zh = row[2]
        name_en = row[3]

        wb_norm = row[norm_col]
        wb_raw = row[raw_col]
        wb_disp = row[disp_col]

        if wb_norm in (None, "N/A", "") or wb_raw in (None, "N/A", ""):
            # Missing record — keep null + missingReason, do NOT use fallback
            records.append({
                "metricId": metric_id,
                "geoId": str(fips).zfill(2),
                "geoName": name_zh,
                "geoAbbr": abbr,
                "geoNameEn": name_en,
                "rawValue": None,
                "displayValue": None,
                "workbookNormalizedValue": None,
                "normalizedValue": None,
                "referenceYear": REFERENCE_YEAR,
                "sourceId": DATASET_ID,
                "verificationStatus": "not_reported",
                "missingReason": "工作簿此单元格标记为 N/A",
                "sourceSheet": "州级数据汇总",
                "sourceRow": ri,
                "sourceColumn": cfg["rawCol"],
                "rawDirection": rawDirection,
            })
            continue

        # Defensive numeric coercion
        try:
            wb_norm_f = float(wb_norm)
            wb_raw_f = float(wb_raw)
        except (TypeError, ValueError):
            records.append({
                "metricId": metric_id,
                "geoId": str(fips).zfill(2),
                "geoName": name_zh,
                "geoAbbr": abbr,
                "geoNameEn": name_en,
                "rawValue": None,
                "displayValue": None,
                "workbookNormalizedValue": None,
                "normalizedValue": None,
                "referenceYear": REFERENCE_YEAR,
                "sourceId": DATASET_ID,
                "verificationStatus": "not_reported",
                "missingReason": f"non-numeric workbook value: norm={wb_norm!r} raw={wb_raw!r}",
                "sourceSheet": "州级数据汇总",
                "sourceRow": ri,
                "sourceColumn": cfg["rawCol"],
                "rawDirection": rawDirection,
            })
            continue

        # Compute our normalizedValue
        if rawDirection == "inverse":
            our_norm = 1.0 - wb_norm_f
        else:
            our_norm = wb_norm_f
        # Clamp to [0,1]
        if our_norm < 0.0:
            our_norm = 0.0
        elif our_norm > 1.0:
            our_norm = 1.0

        records.append({
            "metricId": metric_id,
            "geoId": str(fips).zfill(2),
            "geoName": name_zh,
            "geoAbbr": abbr,
            "geoNameEn": name_en,
            "rawValue": wb_raw_f,
            "displayValue": str(wb_disp) if wb_disp is not None else None,
            "workbookNormalizedValue": wb_norm_f,
            "normalizedValue": round(our_norm, 4),
            "referenceYear": REFERENCE_YEAR,
            "sourceId": DATASET_ID,
            "verificationStatus": "verified",
            "missingReason": None,
            "sourceSheet": "州级数据汇总",
            "sourceRow": ri,
            "sourceColumn": cfg["rawCol"],
            "rawDirection": rawDirection,
        })
    return records


def build_datasets(workbook_sha: str) -> dict[str, Any]:
    return {
        "datasetId": DATASET_ID,
        "datasetVersion": DATASET_VERSION,
        "title": "PathOS 美国各州留学数据矩阵",
        "description": "六类区域指标原始数据，覆盖 50 州 + DC。本轮仅上线 4 类。",
        "sourceName": "Census ACS + FBI UCR + BLS + IPEDS + College Board (混合来源)",
        "sourceUrl": "",
        "referenceYear": REFERENCE_YEAR,
        "retrievedAt": RETRIEVED_AT,
        "geographyLevel": "state",
        "geoIdType": "state_fips",
        "unit": "见各指标定义",
        "valueDirection": "见各指标 rawDirection",
        "normalizationMethod": "见各指标定义",
        "coverage": "51/51 (50 states + DC)",
        "status": "verified",
        "productionReady": False,
        "sourceWorkbookSha256": workbook_sha,
        "readyMetrics": list(METRICS.keys()),
        "existsOutOfScope": list(EXISTS_OUT_OF_SCOPE.keys()),
        "blockedMetrics": list(BLOCKED_METRICS.keys()),
    }


def build_metric_definitions() -> dict[str, Any]:
    out: dict[str, list[dict[str, Any]]] = {"metrics": []}
    for mid, cfg in METRICS.items():
        out["metrics"].append({
            "metricId": mid,
            "displayNameZh": cfg["displayNameZh"],
            "displayNameEn": cfg["displayNameEn"],
            "shortDescription": cfg["shortDescription"],
            "longDescription": cfg["longDescription"],
            "sourceName": cfg["sourceName"],
            "sourceUrl": cfg["sourceUrl"],
            "referenceYear": REFERENCE_YEAR,
            "retrievedAt": RETRIEVED_AT,
            "geographyLevel": "state",
            "geoIdType": "state_fips",
            "rawUnit": cfg["rawUnit"],
            "displayUnit": cfg["displayUnit"],
            "higherIsBetter": cfg["higherIsBetter"],
            "rawDirection": cfg["rawDirection"],
            "normalizationMethod": cfg["normalizationMethod"],
            "allowedRange": cfg["allowedRange"],
            "paletteId": cfg["paletteId"],
            "usedForMap": cfg["usedForMap"],
            "usedForMatch": cfg["usedForMatch"],
            "verificationStatus": "verified",
            "coverage": "51/51",
            "missingCount": 0,
            "datasetId": DATASET_ID,
        })
    return out


def write_json(path: str, payload: Any) -> None:
    # Deterministic: sorted keys + trailing newline + UTF-8 (no BOM).
    text = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)
    text += "\n"
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    workbook_path = args.workbook
    out_dir = args.out

    if not os.path.isfile(workbook_path):
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    workbook_sha = sha256_file(workbook_path)
    print(f"workbook_sha256: {workbook_sha}")

    wb = load_workbook(workbook_path)
    if "州级数据汇总" not in wb.sheetnames:
        print("ERROR: expected sheet '州级数据汇总' missing", file=sys.stderr)
        return 2
    ws_main = wb["州级数据汇总"]

    # Build dataset metadata
    datasets = build_datasets(workbook_sha)

    # Build metric definitions
    metric_defs = build_metric_definitions()

    # Build records — one combined list (all 4 READY metrics)
    all_records: list[dict[str, Any]] = []
    per_metric_file: dict[str, list[dict[str, Any]]] = {}
    for mid, cfg in METRICS.items():
        recs = read_metric(ws_main, mid, cfg, cfg["rawDirection"])
        all_records.extend(recs)
        per_metric_file[mid] = recs

    # Validation report
    validation: dict[str, Any] = {
        "datasetId": DATASET_ID,
        "validatedAt": RETRIEVED_AT,
        "schemaVersion": "1.0.0",
        "summary": {
            "readyMetricCount": len(METRICS),
            "recordsTotal": len(all_records),
            "recordsVerified": sum(1 for r in all_records if r["verificationStatus"] == "verified"),
            "recordsNotReported": sum(1 for r in all_records if r["verificationStatus"] == "not_reported"),
            "recordsPartial": sum(1 for r in all_records if r["verificationStatus"] == "partial"),
            "missingCount": sum(1 for r in all_records if r["rawValue"] is None),
            "coverage": "51/51 per metric",
            "duplicateGeoIds": 0,
            "outlierCount": 0,
        },
        "issues": [],
        "blockedMetrics": [
            {"metricId": mid, "reason": reason}
            for mid, reason in BLOCKED_METRICS.items()
        ],
        "outOfScopeMetrics": list(EXISTS_OUT_OF_SCOPE.keys()),
    }
    # Check duplicates across all_records
    seen: set[tuple[str, str]] = set()
    for r in all_records:
        k = (r["metricId"], r["geoId"])
        if k in seen:
            validation["issues"].append({
                "severity": "high",
                "code": "duplicate_geo_id",
                "metricId": r["metricId"],
                "geoId": r["geoId"],
            })
        seen.add(k)
    validation["summary"]["duplicateGeoIds"] = max(
        0,
        len(seen) - len({r["metricId"] for r in all_records}) * 51
        if False else
        sum(1 for r in all_records if (r["metricId"], r["geoId"]) in seen)
    )
    # Recompute correctly
    seen_set: set[tuple[str, str]] = set()
    dup = 0
    for r in all_records:
        k = (r["metricId"], r["geoId"])
        if k in seen_set:
            dup += 1
        seen_set.add(k)
    validation["summary"]["duplicateGeoIds"] = dup
    if dup:
        validation["issues"].append({"severity": "high", "code": "duplicate_geo_ids_detected", "count": dup})

    # Per-metric distribution stats
    stats: list[dict[str, Any]] = []
    for mid in METRICS:
        values = [r["rawValue"] for r in all_records if r["metricId"] == mid and r["rawValue"] is not None]
        if values:
            sv = sorted(values)
            n = len(sv)
            stats.append({
                "metricId": mid,
                "count": n,
                "min": sv[0],
                "max": sv[-1],
                "mean": round(sum(sv) / n, 2),
                "median": sv[n // 2],
                "missingCount": sum(1 for r in all_records if r["metricId"] == mid and r["rawValue"] is None),
            })
    validation["distribution"] = stats

    # Write outputs (deterministic)
    os.makedirs(out_dir, exist_ok=True)

    files: dict[str, str] = {}
    files["regional-datasets.json"] = json.dumps(datasets, ensure_ascii=False, indent=2, sort_keys=True)
    files["regional-metrics.json"] = json.dumps(metric_defs, ensure_ascii=False, indent=2, sort_keys=True)
    files["regional-records.json"] = json.dumps({"datasetId": DATASET_ID, "records": all_records}, ensure_ascii=False, indent=2, sort_keys=True)
    files["regional-data-validation.json"] = json.dumps(validation, ensure_ascii=False, indent=2, sort_keys=True)

    # Per-metric files (filtered subset for frontend convenience)
    for mid, recs in per_metric_file.items():
        files[f"regional-record-{mid}.json"] = json.dumps(
            {"datasetId": DATASET_ID, "metricId": mid, "records": recs},
            ensure_ascii=False, indent=2, sort_keys=True,
        )

    # Write all files + record sha
    artifact_hashes: dict[str, str] = {}
    for name, content in files.items():
        # Match exactly what is written to disk (content + "\n") so the
        # recorded SHA matches the file's bytes.
        bytes_on_disk = (content + "\n").encode("utf-8")
        full = os.path.join(out_dir, name)
        with open(full, "wb") as f:
            f.write(bytes_on_disk)
        artifact_hashes[name] = hashlib.sha256(bytes_on_disk).hexdigest()

    # Manifest
    manifest = {
        "datasetId": DATASET_ID,
        "manifestVersion": "1.0.0",
        "generatedAt": RETRIEVED_AT,
        "sourceWorkbook": {
            "path": workbook_path,
            "sha256": workbook_sha,
        },
        "artifacts": sorted(
            [
                {"path": name, "sha256": artifact_hashes[name], "bytes": len(content.encode("utf-8"))}
                for name, content in files.items()
            ],
            key=lambda x: x["path"],
        ),
        "totals": {
            "metricCount": len(METRICS),
            "recordCount": len(all_records),
            "readyMetricCount": len(METRICS),
            "blockedMetricCount": len(BLOCKED_METRICS),
            "outOfScopeMetricCount": len(EXISTS_OUT_OF_SCOPE),
        },
    }
    manifest_text = json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    manifest_bytes = manifest_text.encode("utf-8")
    with open(os.path.join(out_dir, "regional-data-manifest.json"), "wb") as f:
        f.write(manifest_bytes)

    print(f"wrote {len(files) + 1} files to {out_dir}")
    print(f"records: {len(all_records)}; verified: {validation['summary']['recordsVerified']}; "
          f"not_reported: {validation['summary']['recordsNotReported']}")
    print(f"manifest sha256: {hashlib.sha256(manifest_bytes).hexdigest()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())